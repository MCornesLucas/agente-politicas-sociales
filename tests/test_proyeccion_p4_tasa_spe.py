"""Tests de P4: guardianes de datos y protocolo de backtest.

Los dos guardianes de este módulo nacieron de la lección del proyecto
hermano (los supuestos sin verificar contra la salida real fallan en
silencio): la suma de tramos debe reproducir el total del indicador 1.1,
y la estructura del archivo B.1.1 del INE debe ser exactamente la
esperada antes de leer una sola celda. Aquí se prueba que ambos
realmente detienen la corrida cuando el dato no cumple.
"""

import numpy as np
import openpyxl
import pandas as pd
import pytest

from politicas_sociales.proyeccion_p4_tasa_spe import (
    TRAMOS_0A17,
    denominador_ine,
    evaluar,
    numerador_0a17,
)


def _csv_inau(tmp_path, total_2020):
    filas = []
    valores = [10, 20, 30, 40, 50]  # tramos 0-17 de 2020: suman 150
    for tramo, valor in zip(TRAMOS_0A17, valores):
        filas.append({"indicador_codigo": "1.1", "apertura": tramo,
                      "anio": 2020, "valor": valor})
    filas.append({"indicador_codigo": "1.1", "apertura": "tramo=18 a 20 años",
                  "anio": 2020, "valor": 25})
    filas.append({"indicador_codigo": "1.1", "apertura": "total",
                  "anio": 2020, "valor": total_2020})
    archivo = tmp_path / "inau_spe_nacional.csv"
    pd.DataFrame(filas).to_csv(archivo, index=False)
    return archivo


def test_numerador_suma_solo_los_tramos_0a17(tmp_path):
    archivo = _csv_inau(tmp_path, total_2020=175)
    serie = numerador_0a17(archivo)
    # El recorte 0-17 excluye el tramo 18-20 (25 casos): 150, no 175.
    assert serie[2020] == 150


def test_guardian_detiene_si_la_suma_de_tramos_no_reproduce_el_total(tmp_path):
    archivo = _csv_inau(tmp_path, total_2020=180)
    with pytest.raises(ValueError, match="suma de tramos"):
        numerador_0a17(archivo)


def _xlsx_ine(tmp_path, etiqueta="Ambos sexos", primera_edad=0):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uruguay"
    # El B.1.1 real trae títulos en las filas 1-4; se escriben para que
    # el libro sintético tenga la misma numeración de filas que el real.
    for fila in range(1, 5):
        ws.cell(row=fila, column=1, value=f"Título fila {fila}")
    ws.cell(row=5, column=2, value=2024)
    ws.cell(row=5, column=3, value=2025)
    ws.cell(row=6, column=1, value=etiqueta)
    for edad in range(18):
        ws.cell(row=7 + edad, column=1, value=primera_edad + edad)
        ws.cell(row=7 + edad, column=2, value=100 + edad)
        ws.cell(row=7 + edad, column=3, value=200 + edad)
    archivo = tmp_path / "b11.xlsx"
    wb.save(archivo)
    return archivo


def test_denominador_suma_las_18_edades_simples(tmp_path):
    archivo = _xlsx_ine(tmp_path)
    pob = denominador_ine([2024, 2025], archivo)
    # 2024: suma de 100..117 = 1953; 2025: suma de 200..217 = 3753.
    assert pob[2024] == 1953
    assert pob[2025] == 3753


def test_guardian_detiene_si_la_estructura_del_b11_cambia(tmp_path):
    archivo = _xlsx_ine(tmp_path, etiqueta="Total")
    with pytest.raises(ValueError, match="estructura"):
        denominador_ine([2024], archivo)


def test_serie_amesetada_no_se_proyecta():
    # El caso real de P4: el numerador se amesetó y ningún candidato
    # supera al ingenuo — el resultado correcto es no publicar proyección.
    serie = np.full(6, 5000.0)
    assert evaluar(serie)["modelo"] == "no_proyectable"


def test_serie_lineal_elige_el_candidato_lineal():
    serie = 1000.0 + 50.0 * np.arange(6)
    veredicto = evaluar(serie)
    assert veredicto["modelo"] == "lineal"
    assert veredicto["backtest"]["lineal"][1] < veredicto["backtest"]["ingenuo"][1]
